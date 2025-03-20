from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook('test.xlsx')
ws = wb['Лист1']

cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G']

for col in cols:
    # Выставляем ширину колонок
    ws.column_dimensions[col].width = 20
    for row in range(1, 16):
        # Выставляем высоту строк
        ws.row_dimensions[row].height = 40

        # Добавляем разделение по разрядам (1000000 поменяется на 1 000 000)
        ws[f'{col}{row}'].value = '{0:,}'.format(ws[f'{col}{row}'].value).replace(',', ' ')

        # Изменяем положение текста в ячейке
        alignment = ws[f'{col}{row}'].alignment.copy(horizontal='center', vertical='center', wrap_text=True)
        ws[f'{col}{row}'].alignment = alignment

        # Заливаем ячейку цветом
        ws[f'{col}{row}'].fill = PatternFill(start_color='EAE7E7', end_color='EAE7E7', fill_type='solid')

wb.save('test.xlsx')
